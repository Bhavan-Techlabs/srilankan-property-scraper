import logging
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from data_processor import COLUMNS

logger = logging.getLogger(__name__)

OUTPUT_DIR = "output"
HEADER_FILL = PatternFill("solid", fgColor="E0E0E0")
HEADER_FONT = Font(bold=True)


def _output_path():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(OUTPUT_DIR, f"srilanka_house_sales_{date_str}.xlsx")


def _get_or_create_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)
    _format_header(ws)
    return ws


def _format_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=False)


def get_existing_data(path, sheet_name):
    """Read existing rows from the sheet as list of dicts."""
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def append_rows(sheet_name, new_rows):
    """Append new_rows (list of lists) to the given sheet, creating/loading the workbook as needed."""
    if not new_rows:
        logger.info("No new rows to append to Excel for %s", sheet_name)
        return

    path = _output_path()
    wb = _get_or_create_workbook(path)
    ws = _get_or_create_sheet(wb, sheet_name)

    for row in new_rows:
        ws.append(row)

    _auto_column_widths(ws)
    wb.save(path)
    logger.info("Appended %d rows to Excel sheet '%s' → %s", len(new_rows), sheet_name, path)


def clear_sheet(sheet_name):
    """Clear all data rows, keeping the header."""
    path = _output_path()
    if not os.path.exists(path):
        return
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for row in list(ws.iter_rows(min_row=2)):
        for cell in row:
            cell.value = None
    wb.save(path)
    logger.info("Cleared Excel sheet: %s", sheet_name)


def _auto_column_widths(ws):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def get_output_path():
    return _output_path()
